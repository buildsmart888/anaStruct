"""
TrussExporter — CSV and PDF report generation.
No GUI dependency; receives model/ss/results and file paths.
"""
from __future__ import annotations

import csv
import io
import math
import os


class TrussExporter:
    """Stateless export helpers."""

    # ── CSV ──────────────────────────────────────────────────────────────────

    @staticmethod
    def export_csv(ss, filepath: str) -> None:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Member", "Force (kN)", "Type"])
            for elem_id in ss.elements:
                r = ss.get_element_results(element_id=elem_id)
                force = r.get("N", 0.0)
                row_type = "Tension" if force > 0 else "Compression"
                w.writerow([f"E{elem_id}", f"{force:.2f}", row_type])

    # ── PDF ──────────────────────────────────────────────────────────────────

    @staticmethod
    def export_pdf(
        model,
        ss,
        analysis_results: list[dict],
        filepath: str,
        scale_support_fn,
        add_dimensions_fn,
        STEEL_GRADES: dict | None = None,
        STEEL_PROFILES: dict | None = None,
        LOAD_COMBINATIONS: dict | None = None,
    ) -> None:
        import matplotlib.pyplot as plt
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            Image as RLImage,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        STEEL_GRADES = STEEL_GRADES or {}
        STEEL_PROFILES = STEEL_PROFILES or {}
        LOAD_COMBINATIONS = LOAD_COMBINATIONS or {}

        # --- Display unit conversion (input always kN / m internally) ---
        _FORCE_TO_KN   = {"kN": 1.0, "N": 0.001, "tf": 9.81, "kgf": 0.00981, "kip": 4.448}
        _LENGTH_TO_M   = {"m": 1.0, "cm": 0.01, "mm": 0.001, "in": 0.0254, "ft": 0.3048}
        disp_fu  = getattr(model, "unit_force",  "kN")
        disp_lu  = getattr(model, "unit_length", "m")
        ff       = _FORCE_TO_KN.get(disp_fu,  1.0)   # divide kN  → display unit
        lf       = _LENGTH_TO_M.get(disp_lu,   1.0)  # divide m   → display unit
        def _F(val_kN):   return val_kN / ff          # kN  → display force
        def _L(val_m):    return val_m  / lf          # m   → display length

        # --- Font ---
        font_name = "Helvetica"
        try:
            tp = r"C:\Windows\Fonts\tahoma.ttf"
            if os.path.exists(tp):
                pdfmetrics.registerFont(TTFont("Tahoma", tp))
                tbp = r"C:\Windows\Fonts\tahomabd.ttf"
                if os.path.exists(tbp):
                    pdfmetrics.registerFont(TTFont("Tahoma-Bold", tbp))
                font_name = "Tahoma"
        except Exception:
            pass
        b_font = "Tahoma-Bold" if font_name == "Tahoma" else "Helvetica-Bold"

        # --- Styles ---
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title", parent=styles["Heading1"], fontName=b_font, fontSize=20,
            textColor=colors.HexColor("#2563EB"), alignment=1, spaceAfter=10,
        )
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"], fontName=b_font, fontSize=14,
            textColor=colors.HexColor("#1F2937"), spaceBefore=15, spaceAfter=8,
        )
        normal = ParagraphStyle("Normal", parent=styles["Normal"], fontName=font_name, fontSize=10)
        small  = ParagraphStyle("Small",  parent=styles["Normal"], fontName=font_name, fontSize=8,
                                textColor=colors.HexColor("#555555"))

        # --- Helper: force light background on any matplotlib figure ---
        def _force_light(fig):
            fig.patch.set_facecolor("#FFFFFF")
            for ax in fig.get_axes():
                ax.set_facecolor("#FFFFFF")
                ax.tick_params(colors="#334155")
                ax.xaxis.label.set_color("#0F172A")
                ax.yaxis.label.set_color("#0F172A")
                ax.title.set_color("#0F172A")
                for spine in ax.spines.values():
                    spine.set_edgecolor("#CBD5E1")
                for txt in ax.texts:
                    if txt.get_color() in ("#FFFFFF", "white", "#ffffff"):
                        txt.set_color("#0F172A")

        # --- Helper: fig → ReportLab image ---
        def fig_to_img(fig, w_ratio=1.0):
            _force_light(fig)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
            buf.seek(0)
            img = RLImage(buf)
            tw = 515 * w_ratio
            img.drawHeight = tw * (img.drawHeight / float(img.drawWidth))
            img.drawWidth = tw
            img.hAlign = "CENTER"
            return img

        # --- Helper: standard table style ---
        def std_table_style(header_color="#D5D8DC"):
            return TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
                ("FONTNAME",   (0, 0), (-1, 0), b_font),
                ("FONTSIZE",   (0, 0), (-1, -1), 9),
                ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.whitesmoke, colors.HexColor("#F0F0F0")]),
            ])

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                rightMargin=40, leftMargin=40,
                                topMargin=40, bottomMargin=40)
        elems = []

        # ════════════════════════════════════════════════════════════════
        # 0. Cover / Header
        # ════════════════════════════════════════════════════════════════
        # ── cover title block ────────────────────────────────────────────
        elems.append(Spacer(1, 60))
        elems.append(Paragraph("STRUCTURAL ANALYSIS REPORT", title_style))
        elems.append(Paragraph(
            "Advanced Truss Analyzer PRO v3.0  —  AISC 360-16",
            ParagraphStyle("Sub", alignment=1, fontSize=11, textColor=colors.gray),
        ))
        elems.append(Spacer(1, 30))

        # project info box
        cover_rows = [
            ["Project",  model.project_data["name"],
             "Date",     model.project_data["date"]],
            ["Engineer", model.project_data["engineer"],
             "Location", model.project_data["location"]],
            ["Client",   model.project_data.get("client", "—"),
             "Code",     model.project_data.get("code", "AISC 360-16")],
            ["Method",   model.design_method,
             "Combination", model.selected_combo],
            ["Force unit", disp_fu,
             "Length unit", disp_lu],
        ]
        ct = Table(cover_rows, colWidths=[70, 180, 70, 180])
        ct.setStyle(TableStyle([
            ("FONTNAME",      (0, 0), (-1, -1), font_name),
            ("FONTNAME",      (0, 0), (0, -1), b_font),
            ("FONTNAME",      (2, 0), (2, -1), b_font),
            ("FONTSIZE",      (0, 0), (-1, -1), 10),
            ("BACKGROUND",    (0, 0), (0, -1), colors.HexColor("#EBF5FB")),
            ("BACKGROUND",    (2, 0), (2, -1), colors.HexColor("#EBF5FB")),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#BDC3C7")),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        elems.append(ct)
        elems.append(Spacer(1, 20))

        # table of contents (static)
        toc_data = [
            ["Section", "Title", "Page"],
            ["1", "Load Summary + Structure Model", "2"],
            ["2", "Axial Force Diagram + Reaction Forces", "3"],
            ["3", "Nodal Displacements + Member Utilization", "4"],
            ["4", "Member Analysis Results", "5"],
            ["5", "Node Coordinates & Profile Reference", "6"],
            ["6", "Detailed Member Calculations", "7+"],
        ]
        toc = Table(toc_data, colWidths=[40, 360, 40])
        toc.setStyle(TableStyle([
            ("FONTNAME",   (0, 0), (-1, 0), b_font),
            ("FONTNAME",   (0, 1), (-1, -1), font_name),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.whitesmoke, colors.HexColor("#EBF5FB")]),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#BDC3C7")),
            ("ALIGN",      (0, 0), (0, -1), "CENTER"),
            ("ALIGN",      (2, 0), (2, -1), "CENTER"),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elems.append(Paragraph("Contents", ParagraphStyle(
            "TOC", fontName=b_font, fontSize=11,
            textColor=colors.HexColor("#1F2937"), spaceBefore=10, spaceAfter=6)))
        elems.append(toc)

        elems.append(PageBreak())   # ── end of cover ──

        # ════════════════════════════════════════════════════════════════
        # 1. Load Summary
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("1. Load Summary", h2_style))

        # Combination factors
        all_combos = {}
        for m, combos in LOAD_COMBINATIONS.items():
            all_combos.update(combos)
        # Include custom combinations from model
        for m, combos in model.custom_combinations.items():
            all_combos.update(combos)

        factors = all_combos.get(model.selected_combo, {})
        if factors:
            factor_rows = [["Case", "Factor", "Description"]]
            case_desc = {"DL": "Dead Load", "LL": "Live Load",
                         "WL": "Wind Load", "SL": "Snow Load"}
            for case, fac in factors.items():
                factor_rows.append([case, f"{fac:.2f}", case_desc.get(case, case)])
            tf = Table(factor_rows, colWidths=[80, 80, 200])
            tf.setStyle(std_table_style("#EAEDED"))
            elems.append(Paragraph(f"Combination: <b>{model.selected_combo}</b>", normal))
            elems.append(Spacer(1, 4))
            elems.append(tf)
            elems.append(Spacer(1, 8))

        # Applied loads table
        load_rows = [[
            "#", "Node",
            f"Fx ({disp_fu})", f"Fy ({disp_fu})",
            "Case", "Factor",
            f"Factored Fx ({disp_fu})", f"Factored Fy ({disp_fu})",
        ]]
        for i, ld in enumerate(model.loads_data):
            fac = factors.get(ld["case"], 0.0)
            load_rows.append([
                str(i + 1), f"N{ld['node_id']}",
                f"{_F(ld['fx']):.3f}", f"{_F(ld['fy']):.3f}", ld["case"],
                f"{fac:.2f}",
                f"{_F(ld['fx'] * fac):.3f}", f"{_F(ld['fy'] * fac):.3f}",
            ])
        tl = Table(load_rows, colWidths=[25, 40, 60, 60, 40, 45, 75, 75])
        tl.setStyle(std_table_style())
        elems.append(tl)

        # ════════════════════════════════════════════════════════════════
        # 2. Structure Model  (with per-case load arrows)
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("2. Structure Model", h2_style))
        _CASE_CLR = {"DL": "#E74C3C", "LL": "#2980B9", "WL": "#27AE60", "SL": "#8E44AD"}
        fig_s = ss.show_structure(show=False)
        scale_support_fn(fig_s)
        add_dimensions_fn(fig_s)
        try:
            ax_s  = fig_s.gca()
            xs_s  = [nd["x"] for nd in model.nodes_data]
            ys_s  = [nd["y"] for nd in model.nodes_data]
            span_s = max(max(xs_s) - min(xs_s), max(ys_s) - min(ys_s), 1.0)
            arr_s  = span_s * 0.13
            lbl_s  = span_s * 0.04
            for ld in model.loads_data:
                if not (1 <= ld["node_id"] <= len(model.nodes_data)):
                    continue
                nd_m   = model.nodes_data[ld["node_id"] - 1]
                xn, yn = nd_m["x"], nd_m["y"]
                clr    = _CASE_CLR.get(ld.get("case", "DL"), "#555")
                fx_s   = _F(ld.get("fx", 0.0))
                fy_s   = _F(ld.get("fy", 0.0))
                if abs(fx_s) > 1e-6:
                    sign = 1 if fx_s > 0 else -1
                    ax_s.annotate("", xy=(xn, yn),
                                  xytext=(xn - sign * arr_s, yn),
                                  arrowprops=dict(arrowstyle="->", color=clr,
                                                  lw=1.8, mutation_scale=12), zorder=5)
                    ax_s.text(xn - sign * (arr_s + lbl_s), yn + lbl_s * 0.6,
                              f"Fx={fx_s:.2f}", fontsize=7, color=clr,
                              ha="center", va="bottom",
                              bbox=dict(boxstyle="round,pad=0.15",
                                        fc="white", ec=clr, alpha=0.85), zorder=6)
                if abs(fy_s) > 1e-6:
                    sign = 1 if fy_s > 0 else -1
                    ax_s.annotate("", xy=(xn, yn),
                                  xytext=(xn, yn - sign * arr_s),
                                  arrowprops=dict(arrowstyle="->", color=clr,
                                                  lw=1.8, mutation_scale=12), zorder=5)
                    ax_s.text(xn + lbl_s * 0.6, yn - sign * (arr_s + lbl_s),
                              f"Fy={fy_s:.2f}", fontsize=7, color=clr,
                              ha="left", va="center",
                              bbox=dict(boxstyle="round,pad=0.15",
                                        fc="white", ec=clr, alpha=0.85), zorder=6)
            from matplotlib.lines import Line2D as _L2D
            present_c = {ld.get("case", "DL") for ld in model.loads_data}
            ax_s.legend(
                handles=[_L2D([0],[0], color=_CASE_CLR.get(c,"#555"), lw=2, label=c)
                         for c in ["DL","LL","WL","SL"] if c in present_c],
                fontsize=8, loc="upper right", title="Load Case")
        except Exception:
            pass
        elems.append(fig_to_img(fig_s, w_ratio=0.95))
        plt.close(fig_s)

        elems.append(PageBreak())   # ── Page 2 → 3 ──

        # ════════════════════════════════════════════════════════════════
        # 3. Axial Force Diagram
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("3. Axial Force Diagram", h2_style))
        fig_a, ax_a = plt.subplots(figsize=(10, 5))
        ax_a.set_facecolor("white")
        try:
            max_f = max((abs(r["force"]) for r in analysis_results), default=1.0) or 1.0
            for i, result in enumerate(analysis_results):
                elem = ss.elements[i + 1]
                f = result["force"]
                color = "#E74C3C" if f > 0.1 else ("#2980B9" if f < -0.1 else "#7F8C8D")
                lw = 2 + int(3 * abs(f) / max_f)
                ax_a.plot([elem["start"][0], elem["end"][0]],
                          [elem["start"][1], elem["end"][1]], color=color, lw=lw)
                mx = (elem["start"][0] + elem["end"][0]) / 2
                my = (elem["start"][1] + elem["end"][1]) / 2
                ax_a.text(mx, my, f"{f:.2f}", fontsize=7, ha="center",
                          bbox=dict(boxstyle="round,pad=0.15", fc="white", ec=color, alpha=0.85))
            ax_a.set_aspect("equal"); ax_a.axis("off")
            from matplotlib.lines import Line2D
            ax_a.legend(handles=[
                Line2D([0], [0], color="#E74C3C", lw=2, label="Tension"),
                Line2D([0], [0], color="#2980B9", lw=2, label="Compression"),
            ], fontsize=8, loc="upper right")
        except Exception:
            pass
        elems.append(fig_to_img(fig_a, w_ratio=0.95))
        plt.close(fig_a)

        # ════════════════════════════════════════════════════════════════
        # 4. Reaction Forces
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("4. Reaction Forces", h2_style))

        try:
            rf = ss.reaction_forces   # {node_id: {'x','y','Fx','Fy'}}

            # 4a. Reaction diagram
            fig_rf, ax_rf = plt.subplots(figsize=(10, 5))
            ax_rf.set_facecolor("white")

            for elem in ss.elements.values():
                ax_rf.plot([elem["start"][0], elem["end"][0]],
                           [elem["start"][1], elem["end"][1]],
                           color="#BDC3C7", lw=2, zorder=1)
            for nid, nd in ss.nodes.items():
                ax_rf.plot(nd["x"], nd["y"], "o", color="#7F8C8D", ms=5, zorder=2)

            xs = [nd["x"] for nd in ss.nodes.values()]
            ys = [nd["y"] for nd in ss.nodes.values()]
            span = max(max(xs) - min(xs), max(ys) - min(ys), 1.0)
            max_rf = max((abs(v["Fx"]) + abs(v["Fy"]) for v in rf.values()), default=1.0) or 1.0
            arrow_scale = span * 0.20

            fixed_len = span * 0.12

            for nid, rn in rf.items():
                x, y = rn["x"], rn["y"]
                rfx, rfy = -rn["Fx"], -rn["Fy"]

                def _arrow(dx, dy, val, clr, lbl_off, _x=x, _y=y):
                    if abs(val) < 0.01:
                        return
                    if dy != 0:
                        # Vertical: arrow body below support node
                        head_xy = (_x, _y)
                        tail_xy = (_x, _y - fixed_len)
                        lx = _x + lbl_off[0]
                        ly = _y - fixed_len + lbl_off[1]
                    else:
                        # Horizontal: fixed length to the side in reaction direction
                        ddx = (1 if dx > 0 else -1) * fixed_len
                        head_xy = (_x + ddx, _y)
                        tail_xy = (_x, _y)
                        lx = _x + ddx + lbl_off[0]
                        ly = _y + lbl_off[1]
                    ax_rf.annotate("", xy=head_xy, xytext=tail_xy,
                                   arrowprops=dict(arrowstyle="->", color=clr,
                                                   lw=2.5, mutation_scale=16), zorder=5)
                    ax_rf.text(lx, ly, f"{_F(abs(val)):.2f} {disp_fu}",
                               fontsize=8, fontweight="bold", color=clr,
                               ha="center", va="center",
                               bbox=dict(boxstyle="round,pad=0.2", fc="white",
                                         ec=clr, alpha=0.85), zorder=6)

                off = span * 0.05
                _arrow(rfx, 0,   rfx, "#E74C3C", (0,  -off))
                _arrow(0,   rfy, rfy, "#1A6BB5", (off,  -off))
                ax_rf.text(x, y - fixed_len - span * 0.07, f"N{nid}",
                           fontsize=8, color="#555", ha="center", zorder=6)

            from matplotlib.lines import Line2D as L2D
            ax_rf.legend(handles=[
                L2D([0], [0], color="#E74C3C", lw=2, label="Fx (Horizontal)"),
                L2D([0], [0], color="#1A6BB5", lw=2, label="Fy (Vertical)"),
            ], fontsize=8, loc="upper right")
            ax_rf.set_aspect("equal"); ax_rf.grid(True, ls="--", alpha=0.3)
            elems.append(fig_to_img(fig_rf))
            plt.close(fig_rf)

            # 4b. Reaction table
            react_rows = [[
                "Node", f"X ({disp_lu})", f"Y ({disp_lu})", "Support",
                f"Fx ({disp_fu})", f"Fy ({disp_fu})", f"Resultant ({disp_fu})",
            ]]
            for nid, rn in rf.items():
                nd_data = model.nodes_data[nid - 1]
                rfx, rfy = -rn["Fx"], -rn["Fy"]
                res = math.sqrt(rfx ** 2 + rfy ** 2)
                react_rows.append([
                    f"N{nid}", f"{_L(rn['x']):.3f}", f"{_L(rn['y']):.3f}",
                    nd_data.get("support", "—"),
                    f"{_F(rfx):.3f}", f"{_F(rfy):.3f}", f"{_F(res):.3f}",
                ])
            tr = Table(react_rows, colWidths=[45, 55, 55, 60, 65, 65, 80])
            tr.setStyle(std_table_style("#D6EAF8"))
            elems.append(Spacer(1, 6))
            elems.append(tr)

        except Exception:
            elems.append(Paragraph("Reaction forces not available.", small))

        elems.append(PageBreak())

        # ════════════════════════════════════════════════════════════════
        # 5. Nodal Displacements (diagram + table)
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("5. Nodal Displacements", h2_style))

        # 5a. Displacement diagram (anastruct built-in, auto-scaled for visibility)
        try:
            fig_disp = ss.show_displacement(show=False)
            fig_disp.set_size_inches(10, 5)
            ax_disp = fig_disp.gca()
            ax_disp.set_title("Displacement Diagram (deformation scaled for visibility)",
                              fontsize=10)
            disps_d = ss.displacements
            for nid, nd in ss.nodes.items():
                d = disps_d.get(nid, {})
                dy_mm = d.get("dy", 0.0) * 1000
                if abs(dy_mm) > 1e-6:
                    ax_disp.text(nd["x"], nd["y"],
                                 f"dy={dy_mm:.2f}mm", fontsize=7,
                                 color="purple", ha="center", va="bottom")
            elems.append(fig_to_img(fig_disp))
            plt.close(fig_disp)
        except Exception:
            elems.append(Paragraph("Displacement diagram not available.", small))

        elems.append(Spacer(1, 8))

        # 5b. Displacement table
        try:
            disps = ss.displacements
            disp_rows = [["Node", "X (m)", "Y (m)", "dx (mm)", "dy (mm)", "δ (mm)"]]
            for nid, nd in ss.nodes.items():
                d = disps.get(nid, {"dx": 0.0, "dy": 0.0})
                dx_mm = d["dx"] * 1000
                dy_mm = d["dy"] * 1000
                delta = math.sqrt(dx_mm ** 2 + dy_mm ** 2)
                disp_rows.append([
                    f"N{nid}", f"{nd['x']:.3f}", f"{nd['y']:.3f}",
                    f"{dx_mm:.4f}", f"{dy_mm:.4f}", f"{delta:.4f}",
                ])
            td = Table(disp_rows, colWidths=[45, 65, 65, 75, 75, 75])
            td.setStyle(std_table_style("#E8DAEF"))

            # Highlight max displacement
            max_delta = max(
                (math.sqrt((disps.get(nid, {}).get("dx", 0)*1000)**2 +
                           (disps.get(nid, {}).get("dy", 0)*1000)**2)
                 for nid in ss.nodes), default=0.0)
            for row_i, (nid, _) in enumerate(ss.nodes.items(), start=1):
                d = disps.get(nid, {"dx": 0.0, "dy": 0.0})
                delta = math.sqrt((d["dx"]*1000)**2 + (d["dy"]*1000)**2)
                if max_delta > 0 and abs(delta - max_delta) < 1e-6:
                    td.setStyle(TableStyle([
                        ("BACKGROUND", (5, row_i), (5, row_i),
                         colors.HexColor("#F9E79F")),
                        ("FONTNAME",   (5, row_i), (5, row_i), b_font),
                    ]))
            elems.append(td)
            elems.append(Paragraph(
                "* Yellow cell = maximum displacement node.", small))
        except Exception:
            elems.append(Paragraph("Displacement data not available.", small))

        elems.append(Spacer(1, 10))

        # ════════════════════════════════════════════════════════════════
        # 6. Member Utilization Diagram
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("6. Member Utilization Diagram", h2_style))
        fig_u, ax_u = plt.subplots(figsize=(10, 5))
        ax_u.set_facecolor("white")
        try:
            for i, result in enumerate(analysis_results):
                elem = ss.elements[i + 1]
                util = result["utilization"]
                color = "#27AE60" if util <= 1.0 else "#E74C3C"
                ax_u.plot([elem["start"][0], elem["end"][0]],
                          [elem["start"][1], elem["end"][1]], color=color,
                          lw=3 + min(util * 2, 5))
                mx = (elem["start"][0] + elem["end"][0]) / 2
                my = (elem["start"][1] + elem["end"][1]) / 2
                ax_u.text(mx, my, f"{util:.2f}", fontsize=8, ha="center",
                          bbox=dict(boxstyle="round,pad=0.15", fc="white",
                                    ec=color, alpha=0.8))
            ax_u.set_aspect("equal"); ax_u.axis("off")
            from matplotlib.patches import Patch as MPatch
            ax_u.legend(handles=[
                MPatch(color="#27AE60", label="OK  (util ≤ 1.0)"),
                MPatch(color="#E74C3C", label="FAIL (util > 1.0)"),
            ], fontsize=8, loc="upper right")
        except Exception:
            pass
        elems.append(fig_to_img(fig_u))
        plt.close(fig_u)

        elems.append(PageBreak())

        # ════════════════════════════════════════════════════════════════
        # 7. Member Analysis Results (full detail)
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("7. Member Analysis Results", h2_style))
        mem_rows = [[
            "Member", "Profile", "Grade", "Area\n(cm²)",
            f"Length\n({disp_lu})", f"Force\n({disp_fu})", "Stress\n(MPa)",
            "Type", "Util.", "Status",
        ]]
        for r in analysis_results:
            prof_data = STEEL_PROFILES.get(r["profile"], {})
            area_str = f"{prof_data.get('Area', '—')}"
            grade_str = prof_data.get("Grade", "—")
            mem_rows.append([
                r["member_id"], r["profile"], grade_str, area_str,
                f"{_L(r['length']):.3f}", f"{_F(r['force']):.3f}",
                f"{r.get('stress', 0.0):.1f}",
                r["type"], f"{r['utilization']:.3f}",
                "OK" if r["status"] == "OK" else "FAIL",
            ])
        tm = Table(mem_rows, colWidths=[42, 100, 38, 38, 42, 48, 48, 65, 38, 38])
        tm.setStyle(std_table_style())
        for i, r in enumerate(analysis_results):
            if r["status"] == "FAIL":
                tm.setStyle(TableStyle([
                    ("TEXTCOLOR", (9, i + 1), (9, i + 1), colors.red),
                    ("FONTNAME",  (9, i + 1), (9, i + 1), b_font),
                ]))
        elems.append(tm)

        elems.append(PageBreak())   # ── Page 5 → 6 ──

        # ════════════════════════════════════════════════════════════════
        # 8. Node Coordinates & Supports
        # ════════════════════════════════════════════════════════════════
        elems.append(Paragraph("8. Node Coordinates & Supports", h2_style))
        node_rows = [["Node", f"X ({disp_lu})", f"Y ({disp_lu})", "Support"]]
        for i, n in enumerate(model.nodes_data):
            node_rows.append([f"N{i+1}", f"{_L(n['x']):.4f}", f"{_L(n['y']):.4f}", n["support"]])
        tn = Table(node_rows, colWidths=[80, 120, 120, 140])
        tn.setStyle(std_table_style("#EAEDED"))
        elems.append(tn)

        elems.append(Spacer(1, 10))

        # ════════════════════════════════════════════════════════════════
        # 9. Steel Profile Reference
        # ════════════════════════════════════════════════════════════════
        used_profiles = list({el["profile"] for el in model.elements_data})
        if used_profiles and STEEL_PROFILES:
            elems.append(Paragraph("9. Steel Profile Reference", h2_style))
            prof_rows = [["Profile", "Grade", "Area (cm²)", "rx (cm)", "ry (cm)",
                          "Fy (MPa)", "E (MPa)"]]
            for pname in sorted(used_profiles):
                pd = STEEL_PROFILES.get(pname, {})
                gd = STEEL_GRADES.get(pd.get("Grade", ""), {})
                prof_rows.append([
                    pname, pd.get("Grade", "—"),
                    f"{pd.get('Area', '—')}",
                    f"{pd.get('rx', '—')}",
                    f"{pd.get('ry', '—')}",
                    f"{gd.get('Fy', '—')}",
                    f"{gd.get('E', '—')}",
                ])
            tp2 = Table(prof_rows, colWidths=[115, 45, 55, 50, 50, 55, 65])
            tp2.setStyle(std_table_style("#FDEBD0"))
            elems.append(tp2)

        elems.append(PageBreak())   # ── Page 6 → 7 ──

        # ════════════════════════════════════════════════════════════════
        # 10. Bill of Materials — Steel Quantity Takeoff
        # ════════════════════════════════════════════════════════════════
        import math as _math

        elems.append(Paragraph("10. Bill of Materials", h2_style))
        elems.append(Spacer(1, 6))

        # --- classify members (same logic as GUI _compute_bom) ---
        _nodes = model.nodes_data
        _elems = model.elements_data
        _ymin  = min(n["y"] for n in _nodes) if _nodes else 0.0
        _htot  = max(n["y"] for n in _nodes) - _ymin if _nodes else 1.0
        _tol   = max(_htot * 0.05, 1e-6)

        def _is_bot(n):
            return n["y"] <= _ymin + _tol

        _groups: dict = {}
        for _el in _elems:
            _na = _nodes[_el["node_a"] - 1]
            _nb = _nodes[_el["node_b"] - 1]
            _dx = abs(_nb["x"] - _na["x"])
            _dy = abs(_nb["y"] - _na["y"])
            _L_m = _math.sqrt(_dx ** 2 + _dy ** 2)
            _pname = _el["profile"]
            if _is_bot(_na) and _is_bot(_nb):
                _cat = "Bottom Chord"
            elif not _is_bot(_na) and not _is_bot(_nb):
                _cat = "Top Chord"
            elif _dx < 1e-6:
                _cat = "Vertical"
            else:
                _cat = "Diagonal"
            _key = (_cat, _pname)
            if _key not in _groups:
                _groups[_key] = {"category": _cat, "profile": _pname,
                                 "count": 0, "total_length": 0.0}
            _groups[_key]["count"] += 1
            _groups[_key]["total_length"] += _L_m

        _cat_order = {"Top Chord": 0, "Bottom Chord": 1, "Vertical": 2, "Diagonal": 3}
        _bom_rows = sorted(_groups.values(), key=lambda r: _cat_order.get(r["category"], 9))
        _CAT_COLOR = {
            "Top Chord":    colors.HexColor("#E74C3C"),
            "Bottom Chord": colors.HexColor("#2980B9"),
            "Vertical":     colors.HexColor("#27AE60"),
            "Diagonal":     colors.HexColor("#F39C12"),
        }

        bom_hdr_row = [
            Paragraph(f"<b>Category</b>",          ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
            Paragraph(f"<b>Profile</b>",           ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
            Paragraph(f"<b>Qty</b>",               ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
            Paragraph(f"<b>Length ({disp_lu})</b>",ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
            Paragraph(f"<b>kg/m</b>",              ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
            Paragraph(f"<b>Weight (kg)</b>",       ParagraphStyle("bh", fontName=b_font, fontSize=8, textColor=colors.white)),
        ]
        bom_data   = [bom_hdr_row]
        _tot_count = 0
        _tot_len   = 0.0
        _tot_wt    = 0.0

        _ps_cell = ParagraphStyle("bc", fontName=font_name, fontSize=8,
                                  textColor=colors.HexColor("#0F172A"))
        for _r in _bom_rows:
            _area    = STEEL_PROFILES.get(_r["profile"], {}).get("Area", 0.0)
            _kgpm    = _area * 0.785
            _wt      = _kgpm * _r["total_length"]
            _disp_l  = _r["total_length"] / lf
            bom_data.append([
                Paragraph(_r["category"], _ps_cell),
                Paragraph(_r["profile"],  _ps_cell),
                str(_r["count"]),
                f"{_disp_l:.2f}",
                f"{_kgpm:.2f}",
                f"{_wt:.1f}",
            ])
            _tot_count += _r["count"]
            _tot_len   += _r["total_length"]
            _tot_wt    += _wt

        # Total row
        _ps_tot = ParagraphStyle("bt", fontName=b_font, fontSize=8, textColor=colors.white)
        bom_data.append([
            Paragraph("<b>TOTAL</b>", _ps_tot), "",
            Paragraph(f"<b>{_tot_count}</b>", _ps_tot),
            Paragraph(f"<b>{_tot_len / lf:.2f}</b>", _ps_tot),
            "",
            Paragraph(f"<b>{_tot_wt:.1f}</b>", _ps_tot),
        ])

        bom_table = Table(bom_data, colWidths=[100, 130, 35, 75, 55, 75])
        bom_style = TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  b_font),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
             [colors.HexColor("#F8FAFC"), colors.HexColor("#E2E8F0")]),
            ("BACKGROUND",   (0, -1), (-1, -1), colors.HexColor("#0F172A")),
            ("TEXTCOLOR",    (0, -1), (-1, -1), colors.white),
            ("FONTNAME",     (0, -1), (-1, -1), b_font),
            ("GRID",         (0, 0),  (-1, -1), 0.5, colors.HexColor("#94A3B8")),
            ("VALIGN",       (0, 0),  (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0),  (-1, -1), 4),
        ])
        # Colour left border of each data row by category
        for _ri, _r in enumerate(_bom_rows, start=1):
            _c = _CAT_COLOR.get(_r["category"], colors.grey)
            bom_style.add("BACKGROUND", (0, _ri), (0, _ri), _c)
            bom_style.add("TEXTCOLOR",  (0, _ri), (0, _ri), colors.white)
        bom_table.setStyle(bom_style)
        elems.append(bom_table)
        elems.append(Spacer(1, 10))

        # Note on density assumption
        note_style = ParagraphStyle("note", fontName=font_name, fontSize=7,
                                    textColor=colors.HexColor("#64748B"))
        elems.append(Paragraph(
            "Note: Steel unit weight = 7 850 kg/m³ (Area × 0.785 × length). "
            "Lengths are gross member lengths — add cutting/splicing allowance as required.",
            note_style))

        elems.append(PageBreak())   # ── BOM → Detailed Calc ──

        # ════════════════════════════════════════════════════════════════
        # 11. Detailed Calculation — Critical Members
        # ════════════════════════════════════════════════════════════════
        dm = getattr(model, "design_method", "LRFD").upper()

        # ── Calculation Sheet header bar ─────────────────────────────────
        calc_hdr_data = [[
            Paragraph(f"<b>MEMBER DESIGN CALCULATIONS</b>", ParagraphStyle(
                "CH", fontName=b_font, fontSize=12, textColor=colors.white)),
            Paragraph(
                f"Project: {model.project_data['name']}  |  "
                f"Method: AISC 360-16 {dm}  |  "
                f"Combo: {model.selected_combo}  |  "
                f"Date: {model.project_data['date']}",
                ParagraphStyle("CHsub", fontName=font_name, fontSize=8, textColor=colors.white)),
        ]]
        calc_hdr = Table([[calc_hdr_data[0][0]], [calc_hdr_data[0][1]]], colWidths=[515])
        calc_hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#1E3A5F")),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ]))
        elems.append(calc_hdr)
        elems.append(Spacer(1, 4))
        elems.append(Paragraph(
            f"Critical member per type (Top Chord / Bottom Chord / Web) — highest utilization.  "
            f"{'φ = 0.90' if dm == 'LRFD' else 'Ω = 1.67'}  |  "
            f"Internal calc units: <b>kN · m · MPa</b>  "
            f"(display: {disp_fu} / {disp_lu})",
            small,
        ))
        elems.append(Spacer(1, 8))

        # AISC factors
        if dm == "LRFD":
            phi_t, phi_c = 0.90, 0.90
        else:
            omega_t, omega_c = 1.67, 1.67

        # --- classify members by geometry ---
        nodes_list = model.nodes_data
        elems_list = model.elements_data
        if nodes_list and elems_list:
            ymin_val = min(n["y"] for n in nodes_list)
            y_tol    = max((max(n["y"] for n in nodes_list) - ymin_val) * 0.05, 1e-6)
            def _member_type_exp(na_idx, nb_idx):
                a_bot = nodes_list[na_idx]["y"] <= ymin_val + y_tol
                b_bot = nodes_list[nb_idx]["y"] <= ymin_val + y_tol
                if a_bot and b_bot:       return "Bottom Chord"
                if not a_bot and not b_bot: return "Top Chord"
                return "Web"

        # --- find critical member per type ---
        type_best: dict = {}
        for r in analysis_results:
            if not elems_list:
                break
            idx   = int(r["member_id"][1:]) - 1
            el    = elems_list[idx]
            mtype = _member_type_exp(el["node_a"] - 1, el["node_b"] - 1)
            if mtype not in type_best or r["utilization"] > type_best[mtype]["utilization"]:
                type_best[mtype] = {**r, "_mtype": mtype, "_elem": el}

        # --- Styles for calculation blocks ---
        mono_style = ParagraphStyle(
            "Mono", parent=styles["Normal"], fontName=font_name,
            fontSize=8, leading=13, leftIndent=16,
        )
        head3_style = ParagraphStyle(
            "Head3", parent=styles["Normal"], fontName=b_font,
            fontSize=11, textColor=colors.HexColor("#1E3A5F"),
            spaceBefore=14, spaceAfter=4,
        )
        step_style = ParagraphStyle(
            "Step", parent=styles["Normal"], fontName=b_font,
            fontSize=9, textColor=colors.HexColor("#2563EB"),
            spaceBefore=6, spaceAfter=2,
        )
        sep_style = ParagraphStyle(
            "Sep", parent=styles["Normal"], fontName=font_name,
            fontSize=8, textColor=colors.HexColor("#AAAAAA"),
        )

        def _hr():
            return Paragraph("─" * 90, sep_style)

        def _box(text, ok):
            bg = "#E8F8F5" if ok else "#FDEDEC"
            ec = "#1E8449" if ok else "#C0392B"
            return Paragraph(
                f'<font color="{ec}"><b>{text}</b></font>',
                ParagraphStyle("Box", parent=styles["Normal"], fontName=b_font,
                               fontSize=10, backColor=colors.HexColor(bg),
                               borderPadding=6, spaceBefore=4, spaceAfter=4),
            )

        for mtype_order in ("Top Chord", "Bottom Chord", "Web"):
            r = type_best.get(mtype_order)
            if r is None:
                continue

            el_data  = r["_elem"]
            prof_nm  = r["profile"]
            prof     = STEEL_PROFILES.get(prof_nm, {})
            grade_nm = prof.get("Grade", "—")
            grade    = STEEL_GRADES.get(grade_nm, {})

            na_d = nodes_list[el_data["node_a"] - 1]
            nb_d = nodes_list[el_data["node_b"] - 1]
            x1, y1 = na_d["x"], na_d["y"]
            x2, y2 = nb_d["x"], nb_d["y"]
            L_m   = r["length"]
            force = r["force"]
            A_cm2 = prof.get("Area", 0.0)
            rx_cm = prof.get("rx", 0.0)
            ry_cm = prof.get("ry", 0.0)
            E_mpa = grade.get("E", 200000.0)
            Fy    = grade.get("Fy", 275.0)
            r_min_cm = min(rx_cm, ry_cm)
            r_min_mm = r_min_cm * 10
            stress = abs(force) * 10.0 / A_cm2 if A_cm2 > 0 else 0.0
            force_type = "TENSION" if force > 0 else "COMPRESSION"
            Pn_kN = Fy * A_cm2 * 0.1   # nominal strength (kN)

            elems.append(Paragraph(
                f"Member {r['member_id']} — {mtype_order}  "
                f"(Critical: highest utilization in this category)",
                head3_style,
            ))
            elems.append(_hr())

            # ── A. Geometry ──────────────────────────────────────────────
            elems.append(Paragraph("A.  Geometry", step_style))
            elems.append(Paragraph(
                f"  Node A = N{el_data['node_a']}   x₁ = {x1:.4f} m,  y₁ = {y1:.4f} m<br/>"
                f"  Node B = N{el_data['node_b']}   x₂ = {x2:.4f} m,  y₂ = {y2:.4f} m<br/>"
                f"<br/>"
                f"  Member length:<br/>"
                f"    L = √[(x₂−x₁)² + (y₂−y₁)²]<br/>"
                f"      = √[({x2:.4f}−{x1:.4f})² + ({y2:.4f}−{y1:.4f})²]<br/>"
                f"      = √[{(x2-x1)**2:.6f} + {(y2-y1)**2:.6f}]<br/>"
                f"      = √{(x2-x1)**2 + (y2-y1)**2:.6f}<br/>"
                f"      = <b>{L_m:.4f} m</b>  =  {L_m*1000:.2f} mm",
                mono_style,
            ))

            # ── B. Section Properties ────────────────────────────────────
            elems.append(Paragraph("B.  Section Properties", step_style))
            elems.append(Paragraph(
                f"  Profile   : {prof_nm}<br/>"
                f"  Steel grade: {grade_nm}<br/>"
                f"  Cross-sectional area:  A  = {A_cm2:.4f} cm²  = {A_cm2*100:.2f} mm²<br/>"
                f"  Radius of gyration:    rx = {rx_cm:.4f} cm  =  {rx_cm*10:.3f} mm<br/>"
                f"                         ry = {ry_cm:.4f} cm  =  {ry_cm*10:.3f} mm<br/>"
                f"  Governing (minimum):   r_min = min(rx, ry) = <b>{r_min_cm:.4f} cm"
                f"  =  {r_min_mm:.3f} mm</b><br/>"
                f"  Modulus of elasticity: E  = {E_mpa:,.0f} MPa<br/>"
                f"  Yield stress:          Fy = {Fy:.0f} MPa",
                mono_style,
            ))

            # ── C. Applied Force ─────────────────────────────────────────
            combo_label = "factored" if dm == "LRFD" else "service (unfactored)"
            force_display_str = (
                f"{force:.4f} kN"
                if disp_fu == "kN"
                else f"{force:.4f} kN  =  {_F(force):.4f} {disp_fu}"
            )
            elems.append(Paragraph("C.  Applied Internal Force  (from FEA)", step_style))
            elems.append(Paragraph(
                f"  N = {force_display_str}   ({force_type}, {combo_label})<br/>"
                f"  Load combination: {model.selected_combo}<br/>"
                f"  Note: AISC calculations below use kN/m/MPa (internal units).",
                mono_style,
            ))

            # ── D. Axial Stress ──────────────────────────────────────────
            elems.append(Paragraph("D.  Axial Stress", step_style))
            elems.append(Paragraph(
                f"  Unit conversion: 1 kN / 1 cm²  =  (1000 N) / (100 mm²)  =  10 N/mm²  =  10 MPa<br/>"
                f"<br/>"
                f"  σ = |N| / A  =  |{force:.4f} kN| × 10 / {A_cm2:.4f} cm²<br/>"
                f"    = {abs(force):.4f} × 10 / {A_cm2:.4f}<br/>"
                f"    = {abs(force)*10:.4f} / {A_cm2:.4f}<br/>"
                f"    = <b>{stress:.4f} MPa</b>",
                mono_style,
            ))

            # ── E. Code Check ────────────────────────────────────────────
            if force > 0:   # TENSION
                elems.append(Paragraph(
                    f"E.  Tension Check — AISC 360-16 §D2  [{dm}]", step_style))

                if dm == "LRFD":
                    design_cap = phi_t * Fy
                    phi_Pn = phi_t * Pn_kN
                    util = stress / design_cap
                    status = "OK" if util <= 1.0 else "FAIL"
                    elems.append(Paragraph(
                        f"  Limit state: Yielding of gross section<br/>"
                        f"  LRFD resistance factor:  φt = {phi_t}<br/>"
                        f"<br/>"
                        f"  Step 1 — Nominal tensile strength:<br/>"
                        f"    Pn = Fy × Ag<br/>"
                        f"       = {Fy:.0f} × {A_cm2:.4f} × 0.1  [kN]<br/>"
                        f"       = <b>{Pn_kN:.4f} kN</b><br/>"
                        f"<br/>"
                        f"  Step 2 — Design tensile strength (LRFD):<br/>"
                        f"    φtPn = φt × Pn<br/>"
                        f"         = {phi_t} × {Pn_kN:.4f}<br/>"
                        f"         = <b>{phi_Pn:.4f} kN</b><br/>"
                        f"<br/>"
                        f"  Equivalent in stress form:<br/>"
                        f"    φt × Fy = {phi_t} × {Fy:.0f} = <b>{design_cap:.2f} MPa</b><br/>"
                        f"<br/>"
                        f"  Step 3 — Check:  Pu ≤ φtPn<br/>"
                        f"    Pu = {abs(force):.4f} kN   vs   φtPn = {phi_Pn:.4f} kN<br/>"
                        f"    or equivalently:  σ ≤ φt×Fy<br/>"
                        f"    {stress:.4f} MPa  vs  {design_cap:.2f} MPa<br/>"
                        f"<br/>"
                        f"  Utilization  U = σ / (φt×Fy)  =  {stress:.4f} / {design_cap:.4f}"
                        f"  =  <b>{util:.4f}</b>",
                        mono_style,
                    ))

                else:  # ASD
                    design_cap = Fy / omega_t
                    allow_Pn = Pn_kN / omega_t
                    util = stress / design_cap
                    status = "OK" if util <= 1.0 else "FAIL"
                    elems.append(Paragraph(
                        f"  Limit state: Yielding of gross section<br/>"
                        f"  ASD safety factor:  Ωt = {omega_t}<br/>"
                        f"<br/>"
                        f"  Step 1 — Nominal tensile strength:<br/>"
                        f"    Pn = Fy × Ag<br/>"
                        f"       = {Fy:.0f} × {A_cm2:.4f} × 0.1  [kN]<br/>"
                        f"       = <b>{Pn_kN:.4f} kN</b><br/>"
                        f"<br/>"
                        f"  Step 2 — Allowable tensile strength (ASD):<br/>"
                        f"    Pn/Ωt = {Pn_kN:.4f} / {omega_t}<br/>"
                        f"          = <b>{allow_Pn:.4f} kN</b><br/>"
                        f"<br/>"
                        f"  Equivalent allowable stress:<br/>"
                        f"    Ft = Fy / Ωt = {Fy:.0f} / {omega_t} = <b>{design_cap:.4f} MPa</b><br/>"
                        f"<br/>"
                        f"  Step 3 — Check:  Pa ≤ Pn/Ωt<br/>"
                        f"    Pa = {abs(force):.4f} kN   vs   Pn/Ωt = {allow_Pn:.4f} kN<br/>"
                        f"    or equivalently:  σ ≤ Ft<br/>"
                        f"    {stress:.4f} MPa  vs  {design_cap:.4f} MPa<br/>"
                        f"<br/>"
                        f"  Utilization  U = σ / Ft  =  {stress:.4f} / {design_cap:.4f}"
                        f"  =  <b>{util:.4f}</b>",
                        mono_style,
                    ))

                elems.append(_box(
                    f"  RESULT:  U = {util:.4f}   {'≤' if util <= 1.0 else '>'} 1.0   →   {status}  ",
                    util <= 1.0,
                ))

            else:   # COMPRESSION
                L_mm = L_m * 1000
                K    = 1.0
                kl_r = K * L_mm / r_min_mm
                Fe   = (math.pi ** 2 * E_mpa) / (kl_r ** 2)
                lim  = 4.71 * math.sqrt(E_mpa / Fy)
                inelastic = kl_r <= lim
                if inelastic:
                    Fcr = (0.658 ** (Fy / Fe)) * Fy
                    mode_str  = "Inelastic buckling (KL/r ≤ 4.71√(E/Fy))"
                    fcr_line1 = f"  Fcr = [0.658^(Fy/Fe)] × Fy"
                    fcr_line2 = f"       = [0.658^({Fy:.0f}/{Fe:.4f})] × {Fy:.0f}"
                    fcr_line3 = f"       = [0.658^{Fy/Fe:.6f}] × {Fy:.0f}"
                    fcr_line4 = f"       = {(0.658**(Fy/Fe)):.6f} × {Fy:.0f}"
                    fcr_line5 = f"       = <b>{Fcr:.4f} MPa</b>"
                else:
                    Fcr = 0.877 * Fe
                    mode_str  = "Elastic buckling (KL/r > 4.71√(E/Fy))"
                    fcr_line1 = f"  Fcr = 0.877 × Fe"
                    fcr_line2 = f"       = 0.877 × {Fe:.4f}"
                    fcr_line3 = f"       = <b>{Fcr:.4f} MPa</b>"
                    fcr_line4 = fcr_line5 = ""

                Pn_c_kN = Fcr * A_cm2 * 0.1   # Pn = Fcr × Ag (kN)
                le_sym  = "≤" if inelastic else ">"

                elems.append(Paragraph(
                    f"E.  Compression Check — AISC 360-16 §E3  [{dm}]", step_style))

                # Steps 1-4 are identical for LRFD and ASD
                elems.append(Paragraph(
                    f"  (Steps 1–4 are common to both LRFD and ASD)<br/>"
                    f"<br/>"
                    f"  Step 1 — Effective length factor:<br/>"
                    f"    K = {K:.1f}  (both ends pin-connected, AISC Commentary §C-C2.2)<br/>"
                    f"    KL = {K:.1f} × {L_mm:.2f} mm  =  {K*L_mm:.2f} mm<br/>"
                    f"<br/>"
                    f"  Step 2 — Slenderness ratio:<br/>"
                    f"    KL/r = KL / r_min<br/>"
                    f"         = {K*L_mm:.2f} / {r_min_mm:.4f}<br/>"
                    f"         = <b>{kl_r:.4f}</b><br/>"
                    f"<br/>"
                    f"  Step 3 — Elastic (Euler) buckling stress:<br/>"
                    f"    Fe = π²E / (KL/r)²<br/>"
                    f"       = π² × {E_mpa:,.0f} / ({kl_r:.4f})²<br/>"
                    f"       = {math.pi**2:.6f} × {E_mpa:,.0f} / {kl_r**2:.4f}<br/>"
                    f"       = {math.pi**2 * E_mpa:.4f} / {kl_r**2:.4f}<br/>"
                    f"       = <b>{Fe:.4f} MPa</b><br/>"
                    f"<br/>"
                    f"  Step 4 — Classify buckling mode:<br/>"
                    f"    Limiting slenderness = 4.71 × √(E/Fy)<br/>"
                    f"                        = 4.71 × √({E_mpa:,.0f} / {Fy:.0f})<br/>"
                    f"                        = 4.71 × √{E_mpa/Fy:.4f}<br/>"
                    f"                        = 4.71 × {math.sqrt(E_mpa/Fy):.4f}<br/>"
                    f"                        = <b>{lim:.4f}</b><br/>"
                    f"    KL/r = {kl_r:.4f}  {le_sym}  {lim:.4f}   →   <b>{mode_str}</b><br/>"
                    f"<br/>"
                    f"  Step 5 — Critical stress (Fcr):<br/>"
                    f"{fcr_line1}<br/>"
                    f"{fcr_line2}<br/>"
                    f"{fcr_line3}<br/>"
                    + (f"{fcr_line4}<br/>{fcr_line5}" if fcr_line4 else ""),
                    mono_style,
                ))

                elems.append(Spacer(1, 4))

                if dm == "LRFD":
                    design_cap = phi_c * Fcr
                    phi_Pn_c  = phi_c * Pn_c_kN
                    util = stress / design_cap
                    status = "OK" if util <= 1.0 else "FAIL"
                    elems.append(Paragraph(
                        f"  Step 6 — Design compressive strength (LRFD):<br/>"
                        f"    φc = {phi_c}  (AISC §E1)<br/>"
                        f"    Nominal strength:   Pn = Fcr × Ag<br/>"
                        f"                           = {Fcr:.4f} × {A_cm2:.4f} × 0.1<br/>"
                        f"                           = <b>{Pn_c_kN:.4f} kN</b><br/>"
                        f"    Design strength:    φcPn = {phi_c} × {Pn_c_kN:.4f}<br/>"
                        f"                             = <b>{phi_Pn_c:.4f} kN</b><br/>"
                        f"<br/>"
                        f"    In stress form: φc × Fcr = {phi_c} × {Fcr:.4f}"
                        f"  =  <b>{design_cap:.4f} MPa</b><br/>"
                        f"<br/>"
                        f"  Step 7 — Check:  Pu ≤ φcPn<br/>"
                        f"    Pu = {abs(force):.4f} kN   vs   φcPn = {phi_Pn_c:.4f} kN<br/>"
                        f"    or: σ ≤ φc×Fcr<br/>"
                        f"    {stress:.4f} MPa  vs  {design_cap:.4f} MPa<br/>"
                        f"<br/>"
                        f"  Utilization  U = σ / (φc×Fcr)"
                        f"  =  {stress:.4f} / {design_cap:.4f}  =  <b>{util:.4f}</b>",
                        mono_style,
                    ))

                else:  # ASD
                    design_cap = Fcr / omega_c
                    allow_Pn_c = Pn_c_kN / omega_c
                    util = stress / design_cap
                    status = "OK" if util <= 1.0 else "FAIL"
                    elems.append(Paragraph(
                        f"  Step 6 — Allowable compressive strength (ASD):<br/>"
                        f"    Ωc = {omega_c}  (AISC §E1)<br/>"
                        f"    Nominal strength:    Pn = Fcr × Ag<br/>"
                        f"                            = {Fcr:.4f} × {A_cm2:.4f} × 0.1<br/>"
                        f"                            = <b>{Pn_c_kN:.4f} kN</b><br/>"
                        f"    Allowable strength:  Pn/Ωc = {Pn_c_kN:.4f} / {omega_c}<br/>"
                        f"                               = <b>{allow_Pn_c:.4f} kN</b><br/>"
                        f"<br/>"
                        f"    In stress form (allowable stress):<br/>"
                        f"    Fa = Fcr / Ωc = {Fcr:.4f} / {omega_c}"
                        f"  =  <b>{design_cap:.4f} MPa</b><br/>"
                        f"<br/>"
                        f"  Step 7 — Check:  Pa ≤ Pn/Ωc<br/>"
                        f"    Pa = {abs(force):.4f} kN   vs   Pn/Ωc = {allow_Pn_c:.4f} kN<br/>"
                        f"    or: σ ≤ Fa<br/>"
                        f"    {stress:.4f} MPa  vs  {design_cap:.4f} MPa<br/>"
                        f"<br/>"
                        f"  Utilization  U = σ / Fa"
                        f"  =  {stress:.4f} / {design_cap:.4f}  =  <b>{util:.4f}</b>",
                        mono_style,
                    ))

                elems.append(_box(
                    f"  RESULT:  U = {util:.4f}   {'≤' if util <= 1.0 else '>'} 1.0   →   {status}  ",
                    util <= 1.0,
                ))

            elems.append(Spacer(1, 16))

        # ════════════════════════════════════════════════════════════════
        # Footer
        # ════════════════════════════════════════════════════════════════
        def footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setFont(font_name, 8)
            canvas.setFillColor(colors.gray)
            canvas.drawRightString(
                A4[0] - 40, 20,
                f"Page {doc_obj.page}  |  Generated by Advanced Truss Analyzer PRO",
            )
            canvas.restoreState()

        doc.build(elems, onFirstPage=footer, onLaterPages=footer)

    # ── Excel ─────────────────────────────────────────────────────────────────

    @staticmethod
    def export_excel(
        model,
        ss,
        analysis_results: list[dict],
        filepath: str,
        all_combo_results: dict | None = None,
        STEEL_GRADES: dict | None = None,
        STEEL_PROFILES: dict | None = None,
        LOAD_COMBINATIONS: dict | None = None,
        scale_support_fn=None,
        add_dimensions_fn=None,
    ) -> None:
        """
        Full Excel report — 7 sheets:
          1. Cover        — project info
          2. Results      — member checks for selected combo
          3. Combo Summary— utilization across all combos (if available)
          4. Diagrams     — Structure / Axial / Displacement / Utilization / Reactions
          5. BOM          — steel quantity takeoff
          6. Nodes        — node coordinates & supports
          7. Profiles     — steel section reference
        """
        import io as _io
        import math as _math
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, GradientFill, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import SeriesLabel

        STEEL_GRADES       = STEEL_GRADES       or {}
        STEEL_PROFILES     = STEEL_PROFILES     or {}
        LOAD_COMBINATIONS  = LOAD_COMBINATIONS  or {}
        all_combo_results  = all_combo_results  or {}

        _FORCE_TO_KN  = {"kN": 1.0, "N": 0.001, "tf": 9.81, "kgf": 0.00981, "kip": 4.448}
        _LENGTH_TO_M  = {"m": 1.0, "cm": 0.01, "mm": 0.001, "in": 0.0254, "ft": 0.3048}
        disp_fu = getattr(model, "unit_force",  "kN")
        disp_lu = getattr(model, "unit_length", "m")
        ff = _FORCE_TO_KN.get(disp_fu, 1.0)
        lf = _LENGTH_TO_M.get(disp_lu, 1.0)
        def _F(v): return v / ff
        def _L(v): return v / lf

        # ── colour palette ───────────────────────────────────────────────────
        C_HEADER_DARK  = "1E3A8A"
        C_HEADER_MID   = "2563EB"
        C_HEADER_LIGHT = "DBEAFE"
        C_OK           = "D1FAE5"
        C_WARN         = "FEF3C7"
        C_FAIL         = "FEE2E2"
        C_OK_TXT       = "065F46"
        C_WARN_TXT     = "92400E"
        C_FAIL_TXT     = "991B1B"
        C_ROW_ALT      = "F8FAFC"
        C_ROW_NORM     = "FFFFFF"
        C_ACCENT       = "06B6D4"
        C_BORDER       = "94A3B8"

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, size=10, color="0F172A", name="Calibri"):
            return Font(name=name, bold=bold, size=size, color=color)

        def _border():
            s = Side(style="thin", color=C_BORDER)
            return Border(left=s, right=s, top=s, bottom=s)

        def _center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _left():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        def _hdr_row(ws, row, values, col_start=1,
                     bg=C_HEADER_DARK, txt="FFFFFF", bold=True, size=10):
            for ci, v in enumerate(values, col_start):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill    = _fill(bg)
                c.font    = _font(bold=bold, size=size, color=txt)
                c.border  = _border()
                c.alignment = _center()

        def _data_row(ws, row, values, col_start=1, alt=False,
                      bold=False, txt="0F172A", fills=None):
            bg = C_ROW_ALT if alt else C_ROW_NORM
            for ci, v in enumerate(values, col_start):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill    = _fill(fills[ci - col_start] if fills else bg)
                c.font    = _font(bold=bold, color=txt)
                c.border  = _border()
                c.alignment = _center()

        def _set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def _merge_title(ws, row, title, n_cols,
                         bg=C_HEADER_DARK, txt="FFFFFF", size=13):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=n_cols)
            c = ws.cell(row=row, column=1, value=title)
            c.fill      = _fill(bg)
            c.font      = _font(bold=True, size=size, color=txt)
            c.alignment = _center()
            ws.row_dimensions[row].height = 24

        def _fig_to_xl_img(fig, width_px=640):
            """Save fig to BytesIO and return openpyxl Image."""
            try:
                import matplotlib
                matplotlib.use("Agg")
            except Exception:
                pass
            buf = _io.BytesIO()
            fig.savefig(buf, format="png", dpi=120, bbox_inches="tight",
                        facecolor="white")
            buf.seek(0)
            img = XLImage(buf)
            ratio = img.height / img.width if img.width else 1
            img.width  = width_px
            img.height = int(width_px * ratio)
            return img

        wb = Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        # ════════════════════════════════════════════════════════════════
        # Sheet 1 — Cover
        # ════════════════════════════════════════════════════════════════
        ws_cov = wb.create_sheet("Cover")
        ws_cov.sheet_view.showGridLines = False
        _set_col_widths(ws_cov, [4, 24, 40, 24, 40, 4])

        # big title
        ws_cov.merge_cells("B2:E3")
        c = ws_cov["B2"]
        c.value     = "STRUCTURAL ANALYSIS REPORT"
        c.font      = _font(bold=True, size=20, color="FFFFFF")
        c.fill      = _fill(C_HEADER_DARK)
        c.alignment = _center()
        ws_cov.row_dimensions[2].height = 30
        ws_cov.row_dimensions[3].height = 30

        ws_cov.merge_cells("B4:E4")
        c = ws_cov["B4"]
        c.value     = "Advanced Truss Analyzer PRO v3.0  —  AISC 360-16"
        c.font      = _font(size=11, color="64748B")
        c.fill      = _fill("EFF6FF")
        c.alignment = _center()
        ws_cov.row_dimensions[4].height = 18

        pd = model.project_data
        info_rows = [
            ("Project",   pd.get("name", "—"),      "Date",      pd.get("date", "—")),
            ("Engineer",  pd.get("engineer", "—"),   "Location",  pd.get("location", "—")),
            ("Client",    pd.get("client", "—"),     "Code",      pd.get("code", "AISC 360-16")),
            ("Method",    model.design_method,        "Combo",     model.selected_combo),
            ("Force unit",disp_fu,                    "Length unit",disp_lu),
        ]
        for ri, (k1, v1, k2, v2) in enumerate(info_rows, start=6):
            ws_cov.row_dimensions[ri].height = 20
            for ci, val in zip([2, 3, 4, 5], [k1, v1, k2, v2]):
                c = ws_cov.cell(row=ri, column=ci, value=val)
                is_key = ci in (2, 4)
                c.font    = _font(bold=is_key, size=11, color="1E3A8A" if is_key else "0F172A")
                c.fill    = _fill("DBEAFE" if is_key else "FFFFFF")
                c.border  = _border()
                c.alignment = _left()

        # TOC
        toc_start = 13
        ws_cov.merge_cells(f"B{toc_start}:E{toc_start}")
        c = ws_cov.cell(row=toc_start, column=2, value="CONTENTS")
        c.font = _font(bold=True, size=12, color="FFFFFF")
        c.fill = _fill(C_HEADER_MID)
        c.alignment = _center()
        toc_items = [
            ("1", "Results",       "Member checks for selected combination"),
            ("2", "Combo Summary", "Utilization across all load combinations"),
            ("3", "Diagrams",      "Structure, Axial, Displacement, Utilization, Reactions"),
            ("4", "BOM",           "Steel quantity takeoff with weights"),
            ("5", "Nodes",         "Node coordinates and support conditions"),
            ("6", "Profiles",      "Steel section reference database"),
        ]
        for ti, (num, sheet, desc) in enumerate(toc_items, toc_start + 1):
            ws_cov.row_dimensions[ti].height = 18
            bg = C_ROW_ALT if ti % 2 == 0 else C_ROW_NORM
            for ci, val in zip([2, 3, 5], [num, sheet, desc]):
                c = ws_cov.cell(row=ti, column=ci, value=val)
                c.font      = _font(size=10, bold=(ci == 3))
                c.fill      = _fill(bg)
                c.border    = _border()
                c.alignment = _left() if ci == 5 else _center()

        # ════════════════════════════════════════════════════════════════
        # Sheet 2 — Results
        # ════════════════════════════════════════════════════════════════
        ws_res = wb.create_sheet("Results")
        ws_res.freeze_panes = "A3"
        _set_col_widths(ws_res, [8, 20, 14, 14, 14, 14, 14, 18, 14, 14, 14])

        _merge_title(ws_res, 1,
            f"Member Analysis Results  —  {model.design_method}  |  {model.selected_combo}  |  "
            f"AISC 360-16",
            11)

        hdrs = ["Member", "Profile", f"Force ({disp_fu})", "Type",
                "Stress (MPa)", "Capacity (MPa)", "Utilization", "Status",
                f"Length ({disp_lu})", "Slenderness", "Design Method"]
        _hdr_row(ws_res, 2, hdrs, bg=C_HEADER_MID)

        for ri, r in enumerate(analysis_results, start=3):
            alt = ri % 2 == 0
            util = r["utilization"]
            if util > 1.0:
                row_bg = C_FAIL
            elif util > 0.8:
                row_bg = C_WARN
            else:
                row_bg = C_OK if alt else C_ROW_NORM

            vals = [
                r["member_id"],
                r["profile"],
                round(_F(r.get("force", 0.0)), 3),
                r.get("type", "—"),
                round(r.get("stress", 0.0), 2),
                round(r.get("design_capacity", 0.0), 2),
                round(util, 4),
                "OK" if r["status"] == "OK" else "FAIL",
                round(_L(r.get("length", 0.0)), 3),
                round(r.get("slenderness", 0.0), 1),
                r.get("design_method", model.design_method),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws_res.cell(row=ri, column=ci, value=val)
                c.border    = _border()
                c.alignment = _center()
                # colour status cell
                if ci == 8:
                    c.fill = _fill(C_FAIL if val == "FAIL" else C_OK)
                    c.font = _font(bold=True,
                                   color=C_FAIL_TXT if val == "FAIL" else C_OK_TXT)
                elif ci == 7:
                    c.font  = _font(bold=True,
                                    color=C_FAIL_TXT if util > 1.0 else
                                    (C_WARN_TXT if util > 0.8 else C_OK_TXT))
                    c.fill  = _fill(row_bg)
                else:
                    c.fill = _fill(row_bg)
                    c.font = _font()

        # summary row
        sr = len(analysis_results) + 3
        ws_res.merge_cells(f"A{sr}:F{sr}")
        total  = len(analysis_results)
        failed = sum(1 for r in analysis_results if r["status"] == "FAIL")
        max_u  = max((r["utilization"] for r in analysis_results), default=0)
        c = ws_res.cell(row=sr, column=1,
            value=f"Total: {total} members  |  Failed: {failed}  |  Max utilization: {max_u:.3f}")
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _fill(C_HEADER_DARK if failed == 0 else "991B1B")
        c.alignment = _center()

        # Utilization bar chart
        if len(analysis_results) > 0:
            chart = BarChart()
            chart.type  = "col"
            chart.title = "Member Utilization"
            chart.y_axis.title = "Utilization ratio"
            chart.x_axis.title = "Member"
            chart.height = 12
            chart.width  = 20
            data_ref = Reference(ws_res,
                                 min_col=7, max_col=7,
                                 min_row=2, max_row=2 + len(analysis_results))
            cats_ref = Reference(ws_res,
                                 min_col=1,
                                 min_row=3, max_row=2 + len(analysis_results))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.solidFill = "2563EB"
            ws_res.add_chart(chart, f"A{sr + 2}")

        # ════════════════════════════════════════════════════════════════
        # Sheet 3 — Combo Summary
        # ════════════════════════════════════════════════════════════════
        ws_cmb = wb.create_sheet("Combo Summary")
        ws_cmb.freeze_panes = "A3"

        combo_names = list(all_combo_results.keys())
        n_combos    = len(combo_names)
        total_cols  = 2 + n_combos + 1   # Member | Profile | combos... | MAX

        _merge_title(ws_cmb, 1,
            f"Utilization Summary — All Load Combinations  ({model.design_method})",
            max(total_cols, 4))
        _set_col_widths(ws_cmb, [10, 22] + [16]*n_combos + [10])

        hdr2 = ["Member", "Profile"] + combo_names + ["MAX"]
        _hdr_row(ws_cmb, 2, hdr2, bg=C_HEADER_MID)

        for ri, r in enumerate(analysis_results, start=3):
            mi = ri - 3
            alt = ri % 2 == 0
            ws_cmb.cell(row=ri, column=1, value=r["member_id"]).border = _border()
            ws_cmb.cell(row=ri, column=2, value=r["profile"]).border   = _border()
            row_max = 0.0
            for ci, cn in enumerate(combo_names, start=3):
                res_list = all_combo_results.get(cn, [])
                u = res_list[mi]["utilization"] if mi < len(res_list) else 0.0
                row_max = max(row_max, u)
                cell = ws_cmb.cell(row=ri, column=ci, value=round(u, 4))
                cell.border    = _border()
                cell.alignment = _center()
                cell.font  = _font(color=C_FAIL_TXT if u > 1.0 else
                                   (C_WARN_TXT if u > 0.8 else C_OK_TXT))
                cell.fill  = _fill(C_FAIL if u > 1.0 else
                                   (C_WARN if u > 0.8 else
                                   (C_ROW_ALT if alt else C_ROW_NORM)))
            mx_cell = ws_cmb.cell(row=ri, column=2 + n_combos + 1,
                                  value=round(row_max, 4))
            mx_cell.font   = _font(bold=True,
                                   color=C_FAIL_TXT if row_max > 1.0 else
                                   (C_WARN_TXT if row_max > 0.8 else C_OK_TXT))
            mx_cell.fill   = _fill(C_FAIL if row_max > 1.0 else
                                   (C_WARN if row_max > 0.8 else C_OK))
            mx_cell.border    = _border()
            mx_cell.alignment = _center()

        if not combo_names:
            ws_cmb.cell(row=3, column=1,
                value="No combo sweep available — run analysis first.").font = _font(size=10)

        # ════════════════════════════════════════════════════════════════
        # Sheet 4 — Diagrams  (matplotlib figures embedded as PNG)
        # ════════════════════════════════════════════════════════════════
        ws_dia = wb.create_sheet("Diagrams")
        ws_dia.sheet_view.showGridLines = False
        _merge_title(ws_dia, 1, "Structural Diagrams", 12)

        def _force_light(fig):
            fig.patch.set_facecolor("#FFFFFF")
            for ax in fig.get_axes():
                ax.set_facecolor("#FFFFFF")
                for spine in ax.spines.values():
                    spine.set_edgecolor("#CBD5E1")

        diagram_row = 3   # current insertion row

        def _place_fig(fig, title, row):
            _force_light(fig)
            ws_dia.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=10)
            c = ws_dia.cell(row=row, column=1, value=title)
            c.font = _font(bold=True, size=12, color="FFFFFF")
            c.fill = _fill(C_HEADER_MID)
            c.alignment = _center()
            ws_dia.row_dimensions[row].height = 20
            img = _fig_to_xl_img(fig, width_px=760)
            anchor_cell = f"A{row + 1}"
            ws_dia.add_image(img, anchor_cell)
            n_rows = max(1, int(img.height / 15)) + 2
            plt.close(fig)
            return row + n_rows

        try:
            fig1 = ss.show_structure(show=False, verbosity=0)
            fig1.suptitle("Structure Model", fontsize=12, fontweight="bold")
            if scale_support_fn:  scale_support_fn(fig1)
            if add_dimensions_fn: add_dimensions_fn(fig1)
            diagram_row = _place_fig(fig1, "1. Structure Model", diagram_row)
        except Exception:
            pass

        try:
            # Axial force diagram
            fig2, ax2 = plt.subplots(figsize=(12, 6))
            elem_ids = list(ss.elements.keys())
            max_f = max([abs(ss.get_element_results(e).get("N", 0)) for e in elem_ids], default=1) or 1
            for eid, elem in ss.elements.items():
                f = ss.get_element_results(eid).get("N", 0)
                clr = "#C0392B" if f > 0.1 else ("#1A5276" if f < -0.1 else "#7F8C8D")
                lw  = 2 + int(3 * abs(f) / max_f)
                ax2.plot([elem["start"][0], elem["end"][0]],
                         [elem["start"][1], elem["end"][1]], color=clr, lw=lw)
                mx, my = (elem["start"][0]+elem["end"][0])/2, (elem["start"][1]+elem["end"][1])/2
                ang = _math.degrees(_math.atan2(elem["end"][1]-elem["start"][1],
                                                elem["end"][0]-elem["start"][0]))
                if abs(ang) > 90: ang += 180
                ax2.text(mx, my, f"{_F(f):.1f}", fontsize=7, ha="center",
                         rotation=ang, color="#0F172A",
                         bbox=dict(boxstyle="round,pad=0.15", fc="white", ec=clr, alpha=0.8))
            ax2.set_aspect("equal"); ax2.grid(True, alpha=0.3)
            ax2.set_title("Axial Force Diagram", fontweight="bold")
            ax2.set_xlabel(f"X ({disp_lu})"); ax2.set_ylabel(f"Y ({disp_lu})")
            diagram_row = _place_fig(fig2, "2. Axial Force Diagram", diagram_row)
        except Exception:
            pass

        try:
            fig3 = ss.show_displacement(show=False)
            fig3.suptitle("Displacement Diagram (Scaled)", fontsize=12, fontweight="bold")
            diagram_row = _place_fig(fig3, "3. Displacement Diagram", diagram_row)
        except Exception:
            pass

        try:
            fig4, ax4 = plt.subplots(figsize=(12, 6))
            if analysis_results:
                for i, r in enumerate(analysis_results):
                    elem = ss.elements[i + 1]
                    u    = r["utilization"]
                    clr  = "#1D8348" if u <= 0.6 else ("#D4AC0D" if u <= 0.9 else
                           ("#E67E22" if u <= 1.0 else "#C0392B"))
                    ax4.plot([elem["start"][0], elem["end"][0]],
                             [elem["start"][1], elem["end"][1]],
                             color=clr, lw=2 + min(u*3, 5))
                    mx = (elem["start"][0]+elem["end"][0])/2
                    my = (elem["start"][1]+elem["end"][1])/2
                    ax4.text(mx, my, f"{u:.2f}", fontsize=7, ha="center",
                             bbox=dict(boxstyle="round,pad=0.15", fc="white", ec=clr, alpha=0.9))
                ax4.set_aspect("equal"); ax4.grid(True, alpha=0.3)
                ax4.set_title("Utilization Diagram", fontweight="bold")
                from matplotlib.patches import Patch
                ax4.legend(handles=[
                    Patch(color="#1D8348", label="≤ 0.6"),
                    Patch(color="#D4AC0D", label="0.6–0.9"),
                    Patch(color="#E67E22", label="0.9–1.0"),
                    Patch(color="#C0392B", label="> 1.0 FAIL"),
                ], loc="upper right", fontsize=8)
            diagram_row = _place_fig(fig4, "4. Member Utilization Diagram", diagram_row)
        except Exception:
            pass

        try:
            rf = ss.reaction_forces
            if rf:
                fig5, ax5 = plt.subplots(figsize=(12, 6))
                for elem in ss.elements.values():
                    ax5.plot([elem["start"][0], elem["end"][0]],
                             [elem["start"][1], elem["end"][1]],
                             color="#94A3B8", lw=1.5)
                xs_r = [n["x"] for n in ss.nodes.values()]
                ys_r = [n["y"] for n in ss.nodes.values()]
                span_r = max(max(xs_r)-min(xs_r), max(ys_r)-min(ys_r), 1.0)
                fl = span_r * 0.12
                for nid, nrf in rf.items():
                    x, y = nrf["x"], nrf["y"]
                    rfx, rfy = -nrf["Fx"], -nrf["Fy"]
                    for dx, dy, val, clr in [(rfx,0,rfx,"#E74C3C"),(0,rfy,rfy,"#2980B9")]:
                        if abs(val) < 0.01: continue
                        if dy != 0:
                            ax5.annotate("", xy=(x,y), xytext=(x, y-fl),
                                arrowprops=dict(arrowstyle="->", color=clr, lw=2))
                            ax5.text(x+span_r*0.03, y-fl+span_r*0.04,
                                     f"{abs(_F(val)):.2f} {disp_fu}",
                                     fontsize=8, color=clr)
                        else:
                            ddx = (dx/abs(dx))*fl
                            ax5.annotate("", xy=(x+ddx,y), xytext=(x,y),
                                arrowprops=dict(arrowstyle="->", color=clr, lw=2))
                            ax5.text(x+ddx+span_r*0.03, y+span_r*0.04,
                                     f"{abs(_F(val)):.2f} {disp_fu}",
                                     fontsize=8, color=clr)
                    ax5.text(x, y-span_r*0.07, f"N{nid}", fontsize=7,
                             ha="center", color="#334155")
                ax5.set_aspect("equal"); ax5.grid(True, alpha=0.3)
                ax5.set_title("Reaction Forces", fontweight="bold")
                diagram_row = _place_fig(fig5, "5. Reaction Forces", diagram_row)
        except Exception:
            pass

        # ════════════════════════════════════════════════════════════════
        # Sheet 5 — BOM
        # ════════════════════════════════════════════════════════════════
        ws_bom = wb.create_sheet("BOM")
        ws_bom.freeze_panes = "A3"
        _set_col_widths(ws_bom, [16, 24, 8, 16, 10, 14])
        _merge_title(ws_bom, 1, "Bill of Materials — Steel Quantity Takeoff", 6)
        _hdr_row(ws_bom, 2,
                 ["Category", "Profile", "Qty", f"Length ({disp_lu})", "kg/m", "Weight (kg)"],
                 bg=C_HEADER_MID)

        nodes_bom  = model.nodes_data
        elems_bom  = model.elements_data
        _ymin_bom  = min(n["y"] for n in nodes_bom) if nodes_bom else 0
        _htot_bom  = max((n["y"] for n in nodes_bom), default=0) - _ymin_bom
        _tol_bom   = max(_htot_bom * 0.05, 1e-6)

        groups: dict = {}
        for el in elems_bom:
            na = nodes_bom[el["node_a"] - 1]
            nb = nodes_bom[el["node_b"] - 1]
            dx = abs(nb["x"] - na["x"])
            dy = abs(nb["y"] - na["y"])
            L  = _math.sqrt(dx**2 + dy**2)
            pn = el["profile"]
            a_bot = na["y"] <= _ymin_bom + _tol_bom
            b_bot = nb["y"] <= _ymin_bom + _tol_bom
            cat = ("Bottom Chord" if a_bot and b_bot else
                   "Top Chord"    if not a_bot and not b_bot else
                   "Vertical"     if dx < 1e-6 else "Diagonal")
            key = (cat, pn)
            if key not in groups:
                groups[key] = {"category": cat, "profile": pn, "count": 0, "total_length": 0.0}
            groups[key]["count"] += 1
            groups[key]["total_length"] += L

        cat_order = {"Top Chord": 0, "Bottom Chord": 1, "Vertical": 2, "Diagonal": 3}
        bom_rows  = sorted(groups.values(), key=lambda r: cat_order.get(r["category"], 9))
        cat_colors = {"Top Chord": "FECACA", "Bottom Chord": "BFDBFE",
                      "Vertical": "BBF7D0", "Diagonal": "FDE68A"}

        tot_count = tot_len = tot_wt = 0.0
        for ri, row in enumerate(bom_rows, start=3):
            alt = ri % 2 == 0
            area   = STEEL_PROFILES.get(row["profile"], {}).get("Area", 0.0)
            kg_pm  = area * 0.785
            tot_wt_row = kg_pm * row["total_length"]
            tot_count += row["count"]
            tot_len   += row["total_length"]
            tot_wt    += tot_wt_row
            cat_bg = cat_colors.get(row["category"], C_ROW_ALT)
            vals = [row["category"], row["profile"], row["count"],
                    round(_L(row["total_length"]), 3),
                    round(kg_pm, 2), round(tot_wt_row, 1)]
            fills = [cat_bg, C_ROW_ALT if alt else C_ROW_NORM,
                     C_ROW_ALT if alt else C_ROW_NORM,
                     C_ROW_ALT if alt else C_ROW_NORM,
                     C_ROW_ALT if alt else C_ROW_NORM,
                     C_ROW_ALT if alt else C_ROW_NORM]
            _data_row(ws_bom, ri, vals, fills=fills)

        # totals row
        tr = len(bom_rows) + 3
        _hdr_row(ws_bom, tr,
                 ["TOTAL", "", int(tot_count),
                  round(_L(tot_len), 3), "", round(tot_wt, 1)],
                 bg=C_HEADER_DARK)
        ws_bom.cell(row=tr+2, column=1,
            value="* Steel unit weight 7 850 kg/m³  (Area cm² × 0.785 × length m)").font = \
            _font(size=8, color="64748B")

        # ════════════════════════════════════════════════════════════════
        # Sheet 6 — Nodes
        # ════════════════════════════════════════════════════════════════
        ws_nod = wb.create_sheet("Nodes")
        ws_nod.freeze_panes = "A3"
        _set_col_widths(ws_nod, [10, 16, 16, 16])
        _merge_title(ws_nod, 1, "Node Coordinates & Boundary Conditions", 4)
        _hdr_row(ws_nod, 2,
                 ["Node", f"X ({disp_lu})", f"Y ({disp_lu})", "Support"],
                 bg=C_HEADER_MID)
        sup_colors = {"Pinned": "DBEAFE", "Roller": "D1FAE5", "Free": C_ROW_NORM}
        for ri, nd in enumerate(model.nodes_data, start=3):
            alt = ri % 2 == 0
            bg = sup_colors.get(nd["support"], C_ROW_ALT)
            _data_row(ws_nod, ri,
                      [f"N{ri-2}", round(_L(nd['x']), 4),
                       round(_L(nd['y']), 4), nd['support']],
                      fills=[C_ROW_ALT if alt else C_ROW_NORM,
                             C_ROW_ALT if alt else C_ROW_NORM,
                             C_ROW_ALT if alt else C_ROW_NORM, bg])

        # ════════════════════════════════════════════════════════════════
        # Sheet 7 — Profiles
        # ════════════════════════════════════════════════════════════════
        ws_pro = wb.create_sheet("Profiles")
        ws_pro.freeze_panes = "A3"
        _set_col_widths(ws_pro, [24, 10, 10, 10, 10, 10, 10, 14, 12])
        _merge_title(ws_pro, 1, "Steel Section Reference", 9)
        _hdr_row(ws_pro, 2,
                 ["Profile", "Area (cm²)", "Ix (cm⁴)", "Iy (cm⁴)",
                  "rx (cm)", "ry (cm)", "kg/m", "Grade", "Standard"],
                 bg=C_HEADER_MID)

        # collect only profiles actually used in this model
        used  = {el["profile"] for el in model.elements_data}
        all_p = STEEL_PROFILES.items()
        # used first, then rest (greyed out)
        sorted_p = sorted(all_p, key=lambda x: (0 if x[0] in used else 1, x[0]))

        for ri, (pname, props) in enumerate(sorted_p, start=3):
            alt   = ri % 2 == 0
            in_use = pname in used
            area  = props.get("Area", 0.0)
            bg    = ("E0F2FE" if in_use else (C_ROW_ALT if alt else C_ROW_NORM))
            from steel_database import SECTION_DB
            std = SECTION_DB.get(pname, {}).get("standard", "—")
            vals = [pname, area, props.get("Ix", 0), props.get("Iy", 0),
                    props.get("rx", 0), props.get("ry", 0),
                    round(area * 0.785, 2),
                    props.get("Grade", "—"), std]
            _data_row(ws_pro, ri, vals,
                      fills=[bg]*9,
                      bold=in_use)

        ws_pro.cell(row=len(list(sorted_p)) + 4, column=1,
                    value="* Highlighted rows = profiles used in this model.").font = \
            _font(size=8, color="64748B")

        wb.save(filepath)
